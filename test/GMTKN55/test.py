import numpy as np
import pandas as pd
import pyscf
from pyscf import dft, lib
import dftd4.pyscf as disp
import torch
from torch.autograd import Variable

import openpyxl
import sys

from gmtkn55 import tools, sets
from snn_3h import SNN


# basic parameters in test
hidden = [40, 40, 40]

use_cuda = False

# fixed basic parameters
lib.num_threads(1)
device = torch.device("cuda" if use_cuda else "cpu")
lamda = 1e-5
beta = 1.5
mindiv = 1e-10
verbose = 9
input_dim = 6
n_net = (
    hidden[0] * (input_dim + 1)
    + hidden[1] * (hidden[0] + 1)
    + hidden[2] * (hidden[1] + 2)
    + 1
)

model_name = "model.log"


def eval_xc_ml(xc_code, rho, weight, coords, spin, relativity=0, deriv=2, verbose=None):
    if spin != 0:
        rho1 = rho[0]
        rho2 = rho[1]
        rho01, dx1, dy1, dz1, lap1, tau1 = rho1[:6]
        rho02, dx2, dy2, dz2, lap2, tau2 = rho2[:6]
        gamma1 = dx1**2 + dy1**2 + dz1**2
        gamma2 = dx2**2 + dy2**2 + dz2**2
        gamma12 = dx1 * dx2 + dy1 * dy2 + dz1 * dz2
    else:
        rho0, dx, dy, dz, lap, tau = rho[:6]
        gamma1 = gamma2 = gamma12 = (dx**2 + dy**2 + dz**2) * 0.25
        rho01 = rho02 = rho0 * 0.5
        tau1 = tau2 = tau * 0.5

    ml_in_ = np.concatenate(
        (
            rho01.reshape((-1, 1)),
            rho02.reshape((-1, 1)),
            gamma1.reshape((-1, 1)),
            gamma12.reshape((-1, 1)),
            gamma2.reshape((-1, 1)),
            tau1.reshape((-1, 1)),
            tau2.reshape((-1, 1)),
        ),
        axis=1,
    )
    ml_in = Variable(torch.Tensor(ml_in_).to(device), requires_grad=True)
    exc_ml_out = s_nn(ml_in)

    ml_exc = exc_ml_out.data[:, 0]
    exc_ml = torch.dot(exc_ml_out[:, 0], ml_in[:, 0] + ml_in[:, 1])
    exc_ml.backward()
    grad = ml_in.grad.cpu().data.numpy()
    grad[np.isnan(grad)] = 0
    grad[np.isinf(grad)] = 0

    if spin != 0:
        vrho_ml = np.hstack((grad[:, 0].reshape((-1, 1)), grad[:, 1].reshape((-1, 1))))
        vgamma_ml = np.hstack(
            (
                grad[:, 2].reshape((-1, 1)),
                grad[:, 3].reshape((-1, 1)),
                grad[:, 4].reshape((-1, 1)),
            )
        )
        vlap_ml = np.zeros((ml_in.shape[0], 2))
        vtau_ml = np.hstack((grad[:, 5].reshape((-1, 1)), grad[:, 6].reshape((-1, 1))))
    else:
        vrho_ml = (grad[:, 0] + grad[:, 1]) / 2
        vgamma_ml = (grad[:, 2] + grad[:, 4] + grad[:, 3]) / 4
        vlap_ml = np.zeros(ml_in.shape[0])
        vtau_ml = (grad[:, 5] + grad[:, 6]) / 2

    # Mix with existing functionals
    dft_xc = dft.libxc.eval_xc("b3lyp", rho, spin, relativity, deriv, verbose)
    dft_exc = np.array(dft_xc[0])
    dft_vrho = np.array(dft_xc[1][0])
    dft_vgamma = np.array(dft_xc[1][1])
    dft_vlap = np.array(dft_xc[1][2])
    dft_vtau = np.array(dft_xc[1][3])

    exc = dft_exc + ml_exc.cpu().numpy()
    vrho = dft_vrho + vrho_ml
    vgamma = dft_vgamma + vgamma_ml
    if dft_vlap == None:
        vlap = vlap_ml
    else:
        vlap = dft_vlap + vlap_ml
    if dft_vtau == None:
        vtau = vtau_ml
    else:
        vtau = dft_vtau + vtau_ml

    vxc = (vrho, vgamma, vlap, vtau)

    return exc, vxc, None, None

def reaction_caculate(reaction_system, file_name):
    reactions = reaction_system.reactions
    sets = reaction_system.systems
    tools.input_set(sets)
    bs = "def2tzvpd"
    vb = 9
    gl = 4
    mc = 200
    atom_num = 0
    names = locals()

    # output file
    # auto set, set_name: subset of gmtkn55
    # file_name: the subset which is caculating
    output = f"/your/path/to/put/this/code/ML-B3LYP-p/test/GMTKN55/data/{file_name}.xlsx"
    wb = openpyxl.Workbook(output)
    wb.save(output)
    wb_read = openpyxl.load_workbook(f"{output}")
    ws = wb_read.active
    count = 2

    # set table's title
    ws.cell(1, 1, "Reaction")
    ws.cell(1, 2, "Reference")
    ws.cell(1, 3, "ML_B3LYP")
    ws.cell(1, 4, "ML_B3LYP+D")
    ws.cell(1, 5, "B3LYP")
    ws.cell(1, 6, "B3LYP + D")
    ws.cell(1, 7, "ML_B3LYP Error")
    ws.cell(1, 8, "ML_B3LYP+D Error")
    ws.cell(1, 9, "B3LYP Error")
    ws.cell(1, 10, "B3LYP + D Error")

    for reaction in reactions:
        print(reaction)
        atom_num = len(reaction["systems"])
        res = 0
        ml_dfa_res = 0
        b3lyp_res = 0
        b3lyp_d_res = 0
        ws.cell(count, 1, str(reaction["systems"]))
        for i in range(atom_num):
            print(reaction["systems"][i])
            names["atom" + str(i)] = tools.getAtom(reaction["systems"][i])
            print(names["atom" + str(i)])
            names["stoichiometry" + str(i)] = reaction["stoichiometry"][i]

            names["mole" + str(i)] = pyscf.M(
                atom=names["atom" + str(i)][0],
                basis=bs,
                verbose=vb,
                spin=names["atom" + str(i)][1],
                charge=names["atom" + str(i)][2],
            )

            print("start mole calculate")
            if names["atom" + str(i)][1] == 0:
                ml_dfa = dft.RKS(names["mole" + str(i)])
                dfa = dft.RKS(names["mole" + str(i)])

            else:
                ml_dfa = dft.UKS(names["mole" + str(i)])
                dfa = dft.UKS(names["mole" + str(i)])

            grid = dft.gen_grid.Grids(names["mole" + str(i)])
            grid.level = gl
            grid.build()

            print("\n\n\n\n\n\n\n\n", lib.num_threads(), "__by azp\n\n\n\n\n\n\n\n")

            # ml-dfa calculate
            ml_dfa = ml_dfa.define_xc_(eval_xc_ml, "MGGA+R", hyb = 0.2)
            ml_dfa.grids = grid
            ml_dfa.max_cycle = mc
            ml_dfa_result = ml_dfa.kernel()

            # b3lyp calculate
            dfa.xc = 'b3lyp'
            dfa.grids = grid
            dfa.max_cycle = mc
            b3lyp_result = dfa.kernel()

            # result of ml-dfa without d
            names["ml_dfa_res" + str(i)] = ml_dfa_result * 627.5095
            ml_dfa_res += names["ml_dfa_res" + str(i)] * int(
                names["stoichiometry" + str(i)]
            )

            # b3lyp result without d
            names["b3lyp_res" + str(i)] = b3lyp_result * 627.5095
            b3lyp_res += names["b3lyp_res" + str(i)] * int(
                names["stoichiometry" + str(i)]
            )

            # dft-D energy
            d4 = disp.DFTD4Dispersion(names["mole" + str(i)], "b3lyp")
            delta_d4 = d4.kernel()[0]

            # b3lyp + d result
            names["res" + str(i)] = (b3lyp_result + delta_d4) * 627.5095
            b3lyp_d_res += names["res" + str(i)] * int(names["stoichiometry" + str(i)])

            # ml-dfa + d result
            names["res" + str(i)] = (ml_dfa_result + delta_d4) * 627.5095
            res += names["res" + str(i)] * int(names["stoichiometry" + str(i)])

        # 参考能量
        ref = reaction["reference"]
        ws.cell(count, 2, ref)

        # 能量值比较
        ws.cell(count, 3, ml_dfa_res)
        ws.cell(count, 4, res)
        ws.cell(count, 5, b3lyp_res)
        ws.cell(count, 6, b3lyp_d_res)

        # 误差分析
        ws.cell(count, 7, abs(ml_dfa_res - ref))
        ws.cell(count, 8, abs(res - ref))
        ws.cell(count, 9, abs(b3lyp_res - ref))
        ws.cell(count, 10, abs(b3lyp_d_res - ref))

        print("b3lyp_res:", b3lyp_res)
        print("b3lyp_d_res:", b3lyp_d_res)
        print("ml_res:", res)
        print("ref:", ref)
        print("---------------")
        count += 1
        wb_read.save(output)

    # pandas mae
    data = pd.read_excel(output)
    ml_b3lyp_mae = data["ML_B3LYP Error"].mean(axis=0)
    ml_b3lyp_d_mae = data["ML_B3LYP+D Error"].mean(axis=0)
    b3lyp_mae = data["B3LYP Error"].mean(axis=0)
    b3lyp_d_mae = data["B3LYP + D Error"].mean(axis=0)
    ws.cell(count, 7, ml_b3lyp_mae)
    ws.cell(count, 8, ml_b3lyp_d_mae)
    ws.cell(count, 9, b3lyp_mae)
    ws.cell(count, 10, b3lyp_d_mae)
    wb_read.save(output)

def process_reaction(set_name):
    reaction_system = getattr(sets, set_name)
    reaction_caculate(reaction_system, set_name)


if __name__ == "__main__":
    # begin calculation
    sets_name = ['W4_11', 'G21EA', 'G21IP', 'DIPCS10', 'PA26', 'SIE4x4', 'ALKBDE10', 'YBDE18', 'AL2X6', 'HEAVYSB11', 'NBPRC', 'ALK8', 'RC21', 'G2RC', 'FH51', 'TAUT15', 'DC13', 'MB16_43', 'DARC',\
                 'RSE43', 'BSR36', 'CDIE20', 'ISO34', 'ISOL24', 'C60ISO', 'PArel', 'BH76', 'BHPERI', 'BHDIV10', 'INV24', 'BHROT27', 'PX13', 'WCPT18', 'RG18', 'ADIM6', 'S22', 'S66', 'HEAVY28',\
                   'WATER27', 'CARBHB12', 'PNICO23', 'HAL59', 'AHB21', 'CHB6', 'IL16', 'IDISP', 'ICONF', 'ACONF', 'Amino20x4', 'PCONF21', 'MCONF', 'SCONF', 'UPU23', 'BUT14DIOL', 'GW100',\
                       'GAPS', 'MRADC', 'ACC24', 'S30L', 'BH9']

    s_nn = SNN(input_dim, hidden, lamda, beta, use_cuda)
    snn_param = s_nn.state_dict()
    s_nn.to(device)

    pp = torch.load(model_name)
    s_nn.load_state_dict(pp)

    set_name = sys.argv[1]
    process_reaction(set_name)
